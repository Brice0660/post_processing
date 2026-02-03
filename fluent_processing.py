
from __future__ import annotations
import numpy as np
import subprocess
import sys
import time
from pathlib import Path
from openpyxl import load_workbook


rows: int = [11,    # Drag Frontwing
             12,    # Drag Sidepod
             13,    # Drag Rearwing
             15,    # Drag Net
             11,    # Df Frontwing
             12,    # Df Sidepod
             13,    # Df Rearwing
             15,    # Df Net
             18     # Moment around front axis
             ]

cols: int = [3,     # Drag Frontwing
             3,     # Drag Sidepod
             3,     # Drag Rearwing
             3,     # Drag Net
             8,     # Df Frontwing
             8,     # Df Sidepod
             8,     # Df Rearwing
             8,     # Df Net
             8      # Moment around front axis
             ]


class FluentPostProcesser():
    def __init__(self, fluent_exe_path: Path, case_file_path: Path):
        self.fluent_exe_path = fluent_exe_path
        self.case_file_path = case_file_path
        self.work_dir = Path(self.case_file_path.parent)
        self.out_dir = self.work_dir / "Processed"
        self.out_dir.mkdir(parents=True, exist_ok=True)
        self.jou_path = self.out_dir / "sequence_30_images.jou"
    
    def create_jou_content(self):
        # --- Parameters of sequence ---
        start_z = 0.0
        end_z = 0.760  # 740 mm
        num_images_y = 30
        num_images_x = 45
        z_positions = np.linspace(start_z, end_z, num_images_y)

        start_x = -1.250
        end_x = 1.850 
        x_positions = np.linspace(start_x, end_x, num_images_x)
        # --- creating jou content ---
        jou_content = f"""/file/read-case-data "{self.case_file_path.as_posix()}"
            /file/confirm-overwrite? no
            /display/set-window 1
            /views/camera/projection orthographic
            /display/display/surface-mesh symmetry()
            /views/auto-scale
            /views/camera/target 0.4 0 1
            /views/camera/position 0.4 1 1
            /views/camera/up-vector 0 0 1
            /views/camera/zoom-camera 9
            """

        for i, z in enumerate(z_positions):
            s_name = f"plane_z_{i:02d}"
            img_vel = f"Processed/Images/side_vel_{i:02d}.png"
            img_pressure = f"Processed/Images/side_pressure_{i:02d}.png"
            
            jou_content += f"""
                ; --- Image {i+1}/{num_images_y} at Z = {z:.4f} ---
                /surface/plane-surface {s_name} z {z:.4f}
                /display/set/contours surfaces {s_name} ()
                /display/contour/velocity-magnitude 0 35   
                /display/save-picture "{img_vel}"
                /display/contour/total-pressure -300 350
                /display/save-picture "{img_pressure}"
                /surface/delete {s_name}
                """

        jou_content += f"""
            views/apply-mirror-planes symmetry ()
            /display/display/surface-mesh Inlet()
            /views/auto-scale
            /views/camera/target 0 0 0.8
            /views/camera/position 1 0 0.8
            /views/camera/up-vector 0 0 1
            /views/camera/zoom-camera 4
            """       

        for i, y in enumerate(x_positions):
            s_name = f"plane_x_{i:02d}"
            img_vel = f"Processed/Images/front_vel_{i:02d}.png"
            img_pressure = f"Processed/Images/front_pressure_{i:02d}.png"
        
            
            jou_content += f"""
                ; --- Image {i+1}/{num_images_x} at x = {y:.4f} ---
                /surface/plane-surface {s_name} y {y:.4f}
                /display/set/contours surfaces {s_name} ()
                /display/contour/velocity-magnitude 0 35
                /display/save-picture "{img_vel}"
                /display/contour/total-pressure -300 300
                /display/save-picture "{img_pressure}"
                /surface/delete {s_name}
                """
            
        jou_content += "\n/exit yes"
        
        self.jou_path.write_text(jou_content, encoding="utf-8")

    def create_images(self, timeout_s = 2000):
        print("Running Fluent (Mode Batch)...")
        self.images_dir = self.out_dir / "Images"
        self.images_dir.mkdir(parents=True, exist_ok=True)
        cmd = [str(self.fluent_exe_path), "3d", "-t8", "-gu", "-i", str(self.jou_path)]
        start = time.time()
        proc = subprocess.Popen(
            cmd,
            cwd=str(self.work_dir),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            universal_newlines=True,
            )
        assert proc.stdout is not None
        try:
            for line in proc.stdout:
                print(line, end="")
                sys.stdout.flush()
                if time.time() - start > timeout_s:
                    proc.kill()
                    raise TimeoutError(f"Process was longer than {timeout_s}s.")
        finally:
            try:
                proc.stdout.close()
            except Exception:
                pass
        rc = proc.wait()
        if rc == 0:
            print(f"\n Images saved in: {self.out_dir}")
        else:
            print(f"\n Error occoured in process: (Code {rc})")

    def get_excel_data(self):
        self.forces = [100, 200, 300, 600, 100, 200, 300, 600, 700]

    def write_to_forcesheet(self):
        self.force_book = load_workbook("aero force sheet.xlsx")
        self.force_sheet = self.force_book.active
        '''
            forces must be in following format:
                [Drag Frontwing,
                 Drag Sidepod,
                 Drag Rearwing,
                 Drag Net,
                 Downforce Frontwing,
                 Downforce Sidepod,
                 Downforce Rearwing,
                 Downforce Net,
                 Moment around front axis]
        '''
        assert len(self.forces) == len(rows) == len(cols)

        assert self.force_sheet["A1"].value == "v1.0"

        for i, row in enumerate(rows):
            self.force_sheet.cell(row=rows[i], column=cols[i], value=self.forces[i])
        
        self.force_book.save("aero force sheet.xlsx")
